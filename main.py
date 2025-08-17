

import os
import json
from typing import Any, Dict, Iterable, List, Optional
import requests
from requests import Session
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from datetime import datetime, timedelta, timezone
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# --------------------------
# Env & constants
# --------------------------
load_dotenv()
OZON_CLIENT_ID = os.getenv("OZON_CLIENT_ID") or os.getenv("CLIENT_ID")
OZON_API_KEY = os.getenv("OZON_API_KEY") or os.getenv("CLIENT_TOKEN") or os.getenv("API_KEY")

API_URL = "https://api-seller.ozon.ru"
DEFAULT_TIMEOUT = 30
MAX_PRODUCT_LIST_PAGE = 1000
MAX_INFO_BATCH = 1000
MAX_STOCK_BATCH = 100
MAX_ORDER_PAGE = 1000
MAX_OFFSET = 20000

class OzonApiError(RuntimeError):
    pass

def _require_env(var_value: Optional[str], var_name: str) -> str:
    if not var_value:
        raise OzonApiError(
            f"Missing required environment variable {var_name}. "
            f"Put it in your .env or shell export."
        )
    return var_value

# --------------------------
# OzonClient
# --------------------------
class OzonClient:
    def __init__(self, client_id: str, api_key: str, base_url: str = API_URL):
        self.base_url = base_url.rstrip("/")
        self.session: Session = requests.Session()
        retry = Retry(total=5, backoff_factor=0.8, status_forcelist=(429,500,502,503,504),
                      allowed_methods=("POST",), raise_on_status=False)
        adapter = HTTPAdapter(max_retries=retry)
        self.session.mount("https://", adapter)
        self.session.mount("http://", adapter)
        self.headers = {"Client-Id": client_id, "Api-Key": api_key, "Content-Type": "application/json"}

    def _post(self, endpoint: str, payload: Dict[str, Any]) -> Any:
        url = f"{self.base_url}/{endpoint.lstrip('/')}"
        resp = self.session.post(url, headers=self.headers, json=payload, timeout=DEFAULT_TIMEOUT)
        try:
            resp.raise_for_status()
        except requests.HTTPError as e:
            try: message = resp.json()
            except: message = resp.text
            raise OzonApiError(f"HTTP error {resp.status_code} on {endpoint}: {message}") from e
        try:
            return resp.json()
        except json.JSONDecodeError as e:
            raise OzonApiError(f"Invalid JSON from {endpoint}: {resp.text[:300]}") from e

    # --- Products / Stocks ---
    def list_clusters(self, cluster_type: str = "CLUSTER_TYPE_OZON") -> List[Dict[str, Any]]:
        data = self._post("v1/cluster/list", {"cluster_type": cluster_type})
        clusters = data.get("clusters") or (data.get("result") or {}).get("clusters", [])
        return clusters or []

    def list_products(self, visibility: str = "IN_SALE", limit: int = MAX_PRODUCT_LIST_PAGE) -> List[Dict[str, Any]]:
        items: List[Dict[str, Any]] = []
        last_id = ""
        while True:
            payload = {"filter": {"visibility": visibility}, "last_id": last_id, "limit": limit}
            data = self._post("v3/product/list", payload)
            result = data.get("result") or {}
            page_items = result.get("items") or []
            items.extend(page_items)
            last_id = result.get("last_id") or ""
            if not last_id: break
        return items

    def get_product_info(self, product_ids: Iterable[int]) -> List[Dict[str, Any]]:
        ids = list(product_ids)
        out: List[Dict[str, Any]] = []
        for i in range(0, len(ids), MAX_INFO_BATCH):
            chunk = ids[i:i+MAX_INFO_BATCH]
            data = self._post("v3/product/info/list", {"product_id": chunk})
            items = data.get("result", {}).get("items") or data.get("items") or []
            out.extend(items if isinstance(items,list) else [])
        return out

    def get_stocks(self, skus: Iterable[int], warehouse_ids: Iterable[int]) -> List[Dict[str, Any]]:
        payload = {"skus": list(skus), "warehouse_ids": list(warehouse_ids)}
        data = self._post("v1/analytics/stocks", payload)
        items = data.get("items") or (data.get("result") or {}).get("items", [])
        return items or []

    # --- Orders ---
    def get_orders(self, since, to) -> List[Dict[str, Any]]:
        orders: List[Dict[str, Any]] = []

        # разбиваем период на интервалы, чтобы не превысить MAX_OFFSET
        current_since = since
        while current_since < to:
            current_to = min(to, current_since + timedelta(days=30))  # берем максимум 1 месяц
            offset = 0

            while True:
                payload = {
                    "dir": "asc",
                    "filter": {
                        "since": current_since.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
                        "to": current_to.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
                    },
                    "limit": MAX_ORDER_PAGE,
                    "offset": offset,
                    "translit": True,
                    "with": {
                        "analytics_data": True,
                        "financial_data": True,
                        "legal_info": False
                    }
                }

                data = self._post("v2/posting/fbo/list", payload)

                # по документации result — это список, а не объект
                result = data.get("result", [])
                if not isinstance(result, list) or not result:
                    break

                orders.extend(result)
                offset += len(result)

                if len(result) < MAX_ORDER_PAGE or offset >= MAX_OFFSET:
                    # если меньше лимита или достигнут max offset, выходим
                    break

            # двигаем период на следующий месяц
            current_since = current_to

        return orders


# --------------------------
# Gather inventory
# --------------------------
def gather_inventory_by_warehouses(client: OzonClient):
    product_list = client.list_products(visibility="IN_SALE")
    product_ids = [p.get("product_id") for p in product_list if "product_id" in p]
    info_items = client.get_product_info(product_ids)
    skus: List[int] = []
    product_meta: Dict[int, Dict[str, Any]] = {}
    for it in info_items:
        sku = it.get("sku")
        if sku is None: continue
        skus.append(int(sku))
        product_meta[int(sku)] = {"name": it.get("name",""), "offer_id": it.get("offer_id","")}

    clusters = client.list_clusters()
    warehouses: List[Dict[str, Any]] = []
    for cl in clusters:
        for lc in cl.get("logistic_clusters", []) or []:
            warehouses.extend(lc.get("warehouses", []) or [])

    inventory: Dict[int, Dict[int, Dict[str,int]]] = {}
    for wh in warehouses:
        wh_id = wh.get("warehouse_id") or wh.get("id")
        if wh_id is None: continue
        for i in range(0, len(skus), MAX_STOCK_BATCH):
            batch = skus[i:i+MAX_STOCK_BATCH]
            stock_items = client.get_stocks(batch, [int(wh_id)])
            for item in stock_items:
                sku_val = item.get("sku")
                if sku_val is None: continue
                inventory.setdefault(int(sku_val), {})[int(wh_id)] = {
                    "available_stock_count": item.get("available_stock_count",0),
                    "transit_stock_count": item.get("transit_stock_count",0),
                    "requested_stock_count": item.get("requested_stock_count",0),
                }
    return inventory, product_meta, clusters

# --------------------------
# Excel export inventory
# --------------------------
def export_inventory_to_excel(product_meta, inventory, clusters, filename="ozon_inventory.xlsx"):
    wb = Workbook()
    wb.remove(wb.active)
    headers = ["Склад", "Артикул", "В наличии", "В пути", "В заявке"]
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Считаем суммарное наличие товаров в кластере
    cluster_sums = []
    for cluster in clusters:
        total_stock = 0
        logistic_clusters = cluster.get("logistic_clusters", [])
        for lc in logistic_clusters:
            for wh in lc.get("warehouses", []):
                wh_id = wh.get("warehouse_id") or wh.get("id")
                for sku in product_meta.keys():
                    total_stock += inventory.get(sku, {}).get(int(wh_id), {}).get("available_stock_count", 0)
        cluster_sums.append((cluster, total_stock))

    # Сортировка кластеров по суммарному наличию (убывание)
    cluster_sums.sort(key=lambda x: x[1], reverse=True)

    for cluster, _ in cluster_sums:
        cluster_name = cluster.get("name", "?")
        ws = wb.create_sheet(title=cluster_name[:31])
        ws.append(headers)
        ws.auto_filter.ref = "A1:E1"
        ws.freeze_panes = "A2"
        ws.sheet_properties.outlinePr.showSummaryBelow = True
        row = 2

        logistic_clusters = cluster.get("logistic_clusters", [])
        for lc in logistic_clusters:
            for wh in lc.get("warehouses", []):
                wh_id = wh.get("warehouse_id") or wh.get("id")
                wh_name = wh.get("name", "?")
                skus_in_wh = [
                    sku for sku in product_meta.keys()
                    if wh_id in inventory.get(sku, {})  # учитываем все товары, даже с нулем
                ]
                if not skus_in_wh:
                    continue

                # Суммируем наличие на складе
                total_stock = sum(inventory.get(sku, {}).get(int(wh_id), {}).get("available_stock_count", 0) for sku in skus_in_wh)

                # Добавляем склад
                ws.append([f"{wh_name} (в наличии: {total_stock})"] + [""] * (len(headers) - 1))
                ws.cell(row=row, column=1).font = Font(bold=True)
                start_row = row + 1
                row += 1

                # Сортируем товары по артикулу (offer_id)
                sorted_skus = sorted(
                    skus_in_wh,
                    key=lambda s: product_meta.get(s, {}).get("offer_id", "")
                )

                for sku in sorted_skus:
                    stock_data = inventory.get(sku, {}).get(int(wh_id), {})
                    ws.append([
                        product_meta.get(sku, {}).get("name", ""),
                        product_meta.get(sku, {}).get("offer_id", ""),
                        stock_data.get("available_stock_count", 0),
                        stock_data.get("transit_stock_count", 0),
                        stock_data.get("requested_stock_count", 0)
                    ])
                    if stock_data.get("available_stock_count", 0) == 0:
                        for c in range(1, 6):
                            ws.cell(row=row, column=c).fill = red_fill
                    row += 1

                # Сгруппируем строки с товарами под складом
                for r in range(start_row, row):
                    ws.row_dimensions[r].outlineLevel = 1
                    ws.row_dimensions[r].hidden = True  # свернуты по умолчанию

    wb.save(filename)
# --------------------------
# Excel export orders summary
# --------------------------
def export_orders_summary_to_excel(client: OzonClient, filename="ozon_last_3_months.xlsx"):
    since = datetime.now() - timedelta(days=90)
    to = datetime.now()
    orders = client.get_orders(since, to)
    summary: Dict[str,int] = {}
    for order in orders:
        for item in order.get("products", []):
            sku = item.get("offer_id") or item.get("sku")
            if sku:
                summary[sku] = summary.get(sku,0) + int(item.get("quantity",1))

    wb = Workbook()
    ws = wb.active
    ws.title = "Заказы 3 месяца"
    ws.append(["Артикул", "Заказы за последние 3 месяца"])
    ws.freeze_panes = "A2"
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Сортируем по артикулу
    for sku in sorted(summary.keys()):
        count = summary[sku]
        ws.append([sku, count])
        if count == 0:
            row = ws.max_row
            ws.cell(row=row, column=2).fill = red_fill

    wb.save(filename)

# --------------------------
# Main
# --------------------------
if __name__ == "__main__":
    client_id = _require_env(OZON_CLIENT_ID, "OZON_CLIENT_ID")
    api_key = _require_env(OZON_API_KEY, "OZON_API_KEY")
    client = OzonClient(client_id, api_key)

    print("Собираем инвентарь...")
    inventory, product_meta, clusters = gather_inventory_by_warehouses(client)
    print("Экспортируем инвентарь в Excel...")
    export_inventory_to_excel(product_meta, inventory, clusters, "ozon_inventory.xlsx")
    print("Экспортируем заказы за последние 3 месяца в Excel...")
    export_orders_summary_to_excel(client, "ozon_last_3_months.xlsx")
    print("Готово! Файлы ozon_inventory.xlsx и ozon_last_3_months.xlsx созданы.")